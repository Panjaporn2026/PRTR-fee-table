import streamlit as st
import pdfplumber
import openpyxl
import tempfile
import shutil
import os
import re

st.set_page_config(page_title="PRTR Fee Table Filler", page_icon="📊", layout="centered")

# ─── Account code → fee category ───────────────────────────────────────────
ACCOUNT_MAP = {
    'E51110101': 'fixed',  'E51110106': 'fixed',  'E51110123': 'fixed',
    'E51110102': 'sso',
    'E51110103': 'pvd',
    'E51110104': 'health_ins',
    'E51110105': 'variable','E51110108': 'variable','E51110109': 'variable',
    'E51110110': 'variable','E51110111': 'variable','E51110112': 'variable',
    'E51110127': 'variable',
    'E51110116': 'variable',   # Expense Refund = variable rate
    'E51110107': 'health_check',
    'E51110124': 'expense',    'E21710101': 'expense',
    'E51110125': 'uniform',
    'E51110126': 'compensation',
    'D51110101': 'fixed',  'D51110106': 'fixed',  'D51110123': 'fixed',
    'D51110104': 'health_ins',
    'D51110105': 'variable','D51110108': 'variable','D51110109': 'variable',
    'D51110110': 'variable','D51110111': 'variable','D51110112': 'variable',
    'D51110127': 'variable',
    'D51110124': 'expense',    'D21710101': 'expense',
    'D51110125': 'uniform',
}


def pad4(lst):
    if not lst:
        return [0.0, 0.0, 0.0, 0.0]
    lst = list(lst)
    while len(lst) < 4:
        lst.append(lst[-1])
    return lst[:4]


def extract_percentages(text):
    pcts = []
    for m in re.finditer(r'(\d+(?:\.\d+)?)\s*%', text):
        v = float(m.group(1))
        if 0 < v <= 100:
            pcts.append(v / 100)
    for m in re.finditer(r'ร้อยละ\s+(\d+(?:\.\d+)?)', text):
        v = float(m.group(1))
        if 0 < v <= 100:
            pcts.append(v / 100)
    return pcts


def extract_all_from_pdf(pdf_bytes):
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
        f.write(pdf_bytes)
        tmp_path = f.name
    try:
        # 1. Try pdfplumber first (fast, works for text-based PDFs)
        text_pages, all_tables = [], []
        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_pages.append(t)
                tbls = page.extract_tables()
                if tbls:
                    all_tables.extend(tbls)
        full_text = '\n'.join(text_pages)

        # 2. Fallback to OCR for scanned PDFs (no embedded text)
        if not full_text.strip():
            try:
                from pdf2image import convert_from_path
                import pytesseract
                images = convert_from_path(tmp_path, dpi=150)
                ocr_parts = []
                for img in images:
                    ocr_parts.append(pytesseract.image_to_string(img, lang='tha+eng'))
                full_text = '\n'.join(ocr_parts)
            except Exception:
                pass  # OCR unavailable — user must enter rates manually

        return full_text, all_tables
    finally:
        os.unlink(tmp_path)


# ─── Company name ────────────────────────────────────────────────────────────
def extract_company_name(text, filename=''):
    patterns = [
        r'(?:และ|And)\s*\n\s*(บริษัท\s+.+?(?:จำกัด|มหาชน)(?:\s*\(มหาชน\))?)',
        r'บริษัท\s+พีอาร์ทีอาร์.{5,200}?\n\s*(บริษัท\s+.+?(?:จำกัด|มหาชน))',
        r'hereinafter referred (?:as|to as) the ["\']Client["\']\.?\s*(?:\n.*?)?(?:and|And)\s+(.+?)\s+whose company',
        r'(?:and|And)\s+([\w\s\(\)\.&\-]+(?:Limited|Co\.,?\s*Ltd\.?|Public Company|PCL|Plc\.?))\s+whose company',
        r'Between\s+PRTR(?:\s+Group)?[^\n]+\s+(?:And|and)\s+(.+?)(?:\n|$)',
        r'(?:And|and)\s*\n\s*(.+?(?:Limited|Ltd\.?|PCL|Co\.))',
        r'"Client"\s+means\s+(.+?(?:Limited|Ltd\.?|PCL))',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
        if m:
            name = re.sub(r'\s+', ' ', m.group(1).strip()).rstrip('.,')
            if 5 < len(name) < 120 and 'PRTR' not in name.upper():
                return name

    if filename:
        clean = re.sub(r'_?OUT[-_]\d+[-_]\d+[-_]\d+[-_R\d]*', '', filename, flags=re.IGNORECASE)
        clean = re.sub(r'_?Completely[\s_]?Signed.*', '', clean, flags=re.IGNORECASE)
        clean = re.sub(r'\.(pdf)$', '', clean, flags=re.IGNORECASE)
        clean = clean.strip('_- ')
        if len(clean) > 3:
            return clean
    return "Unknown Company"


# ─── Category detector (handles garbled Thai) ────────────────────────────────
def detect_cat(col0):
    """
    Detect fee category from table cell text.
    Thai PDFs often garble tone marks; patterns use consonant skeletons.
    e.g. 'รายได้คงที่'  → garbled as 'รายไดค้ งทขี่'
         'รายได้แปรผัน' → garbled as 'รายไดแ้ ปรผนั'  (ปรผ is preserved)
    """
    t = col0

    # Variable income: has "รายได" + "ปรผ" (from แปรผัน)
    if re.search(r'รายได', t, re.IGNORECASE) and re.search(r'ปรผ', t, re.IGNORECASE):
        return 'variable'
    if re.search(r'Variable income', t, re.IGNORECASE):
        return 'variable'

    # Fixed income: has "รายได" but NOT "ปรผ"
    if re.search(r'รายได', t, re.IGNORECASE) and not re.search(r'ปรผ', t, re.IGNORECASE):
        return 'fixed'
    if re.search(r'Fixed income', t, re.IGNORECASE):
        return 'fixed'

    # PVD: เงินกองทนุ สา รองเลยี้ งชพี (garbled) or เงินกองทุนสำรองเลี้ยงชีพ
    if re.search(r'กองทนุ\s*สา\s*รอง|กองทุนส.{0,5}รอง|Provident Fund', t, re.IGNORECASE):
        return 'pvd'

    # Health insurance: ประกนั สขุ ภาพ
    if re.search(r'ประกนั\s*สขุ|ประกันสุข|Health Insurance', t, re.IGNORECASE):
        return 'health_ins'

    # Health check annual (not pre-employment): การตรวจสขุ ภาพประจา ปี
    if re.search(r'ตรวจสขุ.*ประจา|ตรวจสุข.*ประจ|annual.*health|health.*check', t, re.IGNORECASE):
        return 'health_check'

    # Uniform: เครื่องแบบ
    if re.search(r'เครื่องแบบ|Uniform', t, re.IGNORECASE):
        return 'uniform'

    # Expense refund / reimbursement
    if re.search(r'รายจ่าย|ค่าใช.{0,8}ประสาน|Reimbursement|Expense', t, re.IGNORECASE):
        return 'expense'

    # Compensation / severance
    if re.search(r'เงินชดเชย|ค่าชดเชย|Compensation|Severance', t, re.IGNORECASE):
        return 'compensation'

    return None


# ─── Structured table parser ─────────────────────────────────────────────────
def _responsible_party(cols):
    """Return 'PRTR' or 'Client' based on middle columns, or None if unknown."""
    for col in cols[1:-1]:
        if re.search(r'\bPRTR\b|พีอาร์ทีอาร์', col, re.IGNORECASE):
            return 'PRTR'
        if re.search(r'ลูกค้า|\bClient\b', col, re.IGNORECASE):
            return 'Client'
    return None


def parse_fee_table(tables):
    """
    Parse 4-column fee table: หัวข้อ | รายละเอียด | ผู้รับผิดชอบ | ค่าธรรมเนียม
    Tracks continuation rows (col0=None) to accumulate multi-rate categories.
    Returns (rates dict, prtr_cats set) where prtr_cats = categories PRTR pays (no client fee).
    """
    rates = {}
    prtr_cats = set()   # categories where contract says PRTR is responsible & no fee
    fixed_rates = []
    current_cat = None
    current_responsible = None

    for table in tables:
        if not table:
            continue
        for row in table:
            if not row:
                continue
            cols = [str(c or '').strip() for c in row]
            if len(cols) < 2:
                continue

            col0     = cols[0]
            col_rate = cols[-1]   # last column = rate

            # When col0 has content → detect new category
            if col0:
                # Save accumulated fixed rates before switching away
                if current_cat == 'fixed' and fixed_rates and 'fixed' not in rates:
                    rates['fixed'] = fixed_rates[:4] if len(fixed_rates) >= 4 else pad4(fixed_rates)
                    fixed_rates = []

                new_cat = detect_cat(col0)
                current_cat = new_cat
                current_responsible = _responsible_party(cols)
                if new_cat == 'fixed':
                    fixed_rates = []   # reset for new fixed section

            if not current_cat:
                continue

            # Extract % from rate column
            pcts = [float(m) / 100
                    for m in re.findall(r'(\d+(?:\.\d+)?)\s*%', col_rate)
                    if 0 < float(m) <= 100]

            if current_responsible == 'PRTR' and not pcts and current_cat not in rates and current_cat != 'fixed':
                # Contract says PRTR handles this with no client fee
                prtr_cats.add(current_cat)
            elif pcts:
                # Found a real rate → Client pays, remove from prtr_cats if was there
                prtr_cats.discard(current_cat)
                if current_cat == 'fixed':
                    fixed_rates.extend(pcts)
                elif current_cat not in rates:
                    rates[current_cat] = pad4(pcts)

    # Flush remaining fixed rates
    n_fixed_raw = len(fixed_rates)   # count BEFORE padding
    if fixed_rates and 'fixed' not in rates:
        rates['fixed'] = fixed_rates[:4] if len(fixed_rates) >= 4 else pad4(fixed_rates)
    elif 'fixed' in rates:
        # already saved mid-loop; count was set when flushed, approximate from padded list
        # detect repetition: [20,15,15,15] → 2 raw; [18,8,15,7] → 4 raw; [22,22,22,22] → 1 raw
        r = rates['fixed']
        if r[0] == r[1] == r[2] == r[3]:
            n_fixed_raw = 1
        elif r[1] == r[2] == r[3]:
            n_fixed_raw = 2 if r[0] != r[1] else 1
        else:
            n_fixed_raw = 4

    return rates, prtr_cats, n_fixed_raw


# ─── Text-based fallback ──────────────────────────────────────────────────────
def find_rate_near(text, *keywords, default=None):
    for kw in keywords:
        m = re.search(kw + r'[^\n%]{0,120}?(\d+(?:\.\d+)?)\s*%', text, re.IGNORECASE | re.DOTALL)
        if m:
            return float(m.group(1)) / 100
    return default


def extract_text_fallback(text):
    """Text-based rate extraction when table parsing fails."""
    rates = {}
    fixed_sec = re.search(
        r'(?:Fixed income|รายได.{0,8}คงที).{0,600}?(?=รายได.{0,8}แปรผ|Variable income|$)',
        text, re.DOTALL | re.IGNORECASE)
    var_sec = re.search(
        r'(?:Variable income|รายได.{0,8}(?:แปรผ|ปรผ)).{0,400}?(?=รายจ่าย|Expense|กองทนุ|Social Security|$)',
        text, re.DOTALL | re.IGNORECASE)
    fp = extract_percentages(fixed_sec.group(0)) if fixed_sec else []
    vp = extract_percentages(var_sec.group(0)) if var_sec else []
    if fp:
        rates['fixed'] = fp[:4] if len(fp) >= 4 else (fp[:2] if len(fp) >= 2 else pad4(fp))
    if vp:
        rates['variable'] = pad4(vp[:4]) if len(vp) >= 4 else (vp[:2] if len(vp) >= 2 else pad4(vp))
    return rates


# ─── Common rates filler ──────────────────────────────────────────────────────
def fill_common(rates, text, prtr_cats=None):
    """Fill defaults for categories not yet extracted. Skip PRTR-managed categories."""
    if prtr_cats is None:
        prtr_cats = set()

    def get(keys, default):
        return find_rate_near(text, *keys, default=default)

    if 'pvd' not in rates and 'pvd' not in prtr_cats:
        rates['pvd'] = pad4([get([r'Provident Fund', r'กองทนุ\s*สา\s*รอง', r'กองทุนส.{0,5}รอง'], 0.05)])
    if 'health_ins' not in rates and 'health_ins' not in prtr_cats:
        rates['health_ins'] = pad4([get([r'Health Insurance', r'ประกนั\s*สขุ', r'ประกันสุข'], 0.05)])
    if 'health_check' not in rates and 'health_check' not in prtr_cats:
        rates['health_check'] = pad4([get([r'Health Check', r'ตรวจสขุ.*ประจา', r'ตรวจสุข.*ประจ'], 0.05)])
    if 'uniform' not in rates and 'uniform' not in prtr_cats:
        rates['uniform'] = pad4([get([r'Uniform', r'เครื่องแบบ'], 0.05)])
    if 'expense' not in rates and 'expense' not in prtr_cats:
        rates['expense'] = pad4([get([r'Reimbursement', r'ค่าใช.{0,8}ประสาน'], 0.10)])
    if 'sso' not in rates:
        rates['sso'] = [0.0, 0.0, 0.0, 0.0]
    if 'compensation' not in rates and 'compensation' not in prtr_cats:
        cm = re.search(
            r'(?:compensation|ค่าชดเชย|เงินชดเชย).{0,300}?(?:plus|บวก)\s+(\d+(?:\.\d+)?)\s*%',
            text, re.IGNORECASE | re.DOTALL)
        rates['compensation'] = pad4([float(cm.group(1)) / 100 if cm else 0.03])

    for k in rates:
        if isinstance(rates[k], list):
            rates[k] = pad4(rates[k])
    return rates


# ─── Main detection ───────────────────────────────────────────────────────────
def detect_fee_structure(text, tables):
    m = re.search(
        r'(?:ADDENDUM OF THE CONTRACT 1|บันทึกแนบท้ายสัญญา\s*1).*',
        text, re.DOTALL | re.IGNORECASE)
    add = m.group(0) if m else text

    # 1. Structured table parser (most reliable)
    rates, prtr_cats, n_fixed_raw = parse_fee_table(tables)

    # 2. Text fallback for missing categories
    if 'fixed' not in rates or 'variable' not in rates:
        for k, v in extract_text_fallback(add).items():
            if k not in rates:
                rates[k] = v

    # 3. Determine number of fill columns from contract
    if n_fixed_raw >= 4:
        n_cols = 4
    elif n_fixed_raw >= 2:
        n_cols = 2
    else:
        n_cols = 1

    if 'fixed' not in rates:
        rates['fixed'] = [0.20, 0.15]
        n_cols = 2
    if 'variable' not in rates:
        rates['variable'] = rates.get('fixed', [0.15, 0.15])[:n_cols]

    # Recalculate n_cols from actual rates (handles scanned PDF where table=empty, text fallback ran)
    fixed_r = pad4(rates.get('fixed', [0.20]))
    if fixed_r[0] == fixed_r[1] == fixed_r[2] == fixed_r[3]:
        detected_n = 1
    elif fixed_r[1] == fixed_r[2] == fixed_r[3]:
        detected_n = 2 if fixed_r[0] != fixed_r[1] else 1
    else:
        detected_n = 4
    n_cols = max(n_cols, detected_n)

    # 4. Fill remaining common rates (skip PRTR-managed)
    rates = fill_common(rates, add, prtr_cats)
    return n_cols, rates, prtr_cats


# ─── Excel filler ─────────────────────────────────────────────────────────────
def fill_excel(template_bytes, rates, n_cols, prtr_cats=None):
    """
    n_cols: 1=fill J only, 2=fill J+K, 4=fill J+K+L+M
    prtr_cats: set of categories where contract says PRTR is responsible (fill 'PRTR' text)
    """
    if prtr_cats is None:
        prtr_cats = set()

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(template_bytes)
        tmp_in = f.name
    tmp_out = tmp_in + '_out.xlsx'
    shutil.copy2(tmp_in, tmp_out)

    wb = openpyxl.load_workbook(tmp_out)
    ws = wb['3. Form Fee tabel 22.05.26']
    pct_fmt = '0.0%'

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        cost    = row[8].value
        account = str(row[4].value).strip() if row[4].value else ''
        paycode = str(row[6].value).strip() if row[6].value else ''
        header  = str(row[3].value).strip() if row[3].value else ''
        j, k, l, m = row[9], row[10], row[11], row[12]
        all_cells = [j, k, l, m]

        if cost == 'PRTR':
            for cell in all_cells:
                cell.value = 'PRTR'
        elif cost == 'CLNT':
            if paycode == '1 M NOTICE' or header == 'Severance Pay':
                category = 'compensation'
            else:
                category = ACCOUNT_MAP.get(account)
            if category:
                if category in prtr_cats:
                    # Contract specifies PRTR handles this — fill 'PRTR' for active cols, clear rest
                    for i, cell in enumerate(all_cells):
                        cell.value = 'PRTR' if i < n_cols else None
                else:
                    r = pad4(rates.get(category, [0.0]))
                    for i, cell in enumerate(all_cells):
                        if i < n_cols:
                            cell.value = r[i]
                            cell.number_format = pct_fmt
                        else:
                            cell.value = None

    wb.save(tmp_out)
    with open(tmp_out, 'rb') as f:
        out_bytes = f.read()
    os.unlink(tmp_in)
    os.unlink(tmp_out)
    return out_bytes


# ─── UI ───────────────────────────────────────────────────────────────────────
st.title("📊 PRTR Fee Table Filler")
st.caption("อัปโหลดสัญญา PDF + Master Fee Table แล้วระบบจะกรอก fee rate ให้อัตโนมัติ")
st.divider()

col1, col2 = st.columns(2)
with col1:
    pdf_file = st.file_uploader("📄 สัญญา PDF", type=['pdf'],
                                 help="Labor Supplier Contract (OUT-X-XXXX-XXXXX)")
with col2:
    xlsx_file = st.file_uploader("📋 Master Fee Table", type=['xlsx'],
                                  help="Master - Fee tebla.xlsx")

if pdf_file and xlsx_file:
    with st.spinner("กำลังอ่านสัญญา..."):
        pdf_bytes = pdf_file.read()
        full_text, tables = extract_all_from_pdf(pdf_bytes)
        company_name = extract_company_name(full_text, pdf_file.name)
        n_cols, rates, prtr_cats = detect_fee_structure(full_text, tables)

    st.success("✅ อ่านสัญญาเสร็จแล้ว")
    st.subheader("📋 ข้อมูลที่อ่านได้จากสัญญา")
    st.write(f"**บริษัท Client:** {company_name}")

    col_labels_active = ['J', 'K', 'L', 'M'][:n_cols]
    struct_label = {1: '1 Condition (J only)', 2: '2 Conditions (J, K)', 4: '4 Conditions / Tiered (J, K, L, M)'}.get(n_cols, f'{n_cols} Conditions')
    st.write(f"**จำนวน Conditions:** {struct_label}")

    if prtr_cats:
        cat_name_map = {
            'pvd': 'PVD', 'health_ins': 'Health Insurance', 'health_check': 'Health Check',
            'uniform': 'Uniform', 'expense': 'Expense/Reimbursement', 'compensation': 'Compensation',
            'sso': 'SSO',
        }
        names = ', '.join(cat_name_map.get(c, c) for c in sorted(prtr_cats))
        st.info(f"📌 รายการที่สัญญาระบุว่า **PRTR** รับผิดชอบ (จะใส่ 'PRTR' ในไฟล์): {names}")

    import pandas as pd

    def fmt(cat):
        if cat in prtr_cats:
            return ['PRTR'] * n_cols
        return [f"{v*100:.1f}%" for v in rates.get(cat, [0]*4)[:n_cols]]

    rate_display = {
        'Fixed income':           fmt('fixed'),
        'Variable income':        fmt('variable'),
        'Expense Refund':         fmt('variable'),
        'Reimbursement':          fmt('expense'),
        'PVD':                    fmt('pvd'),
        'Health Insurance':       fmt('health_ins'),
        'Severance/Compensation': fmt('compensation'),
    }
    df = pd.DataFrame(rate_display, index=col_labels_active).T
    st.dataframe(df, use_container_width=True)

    with st.expander("🔍 ดูข้อความที่อ่านจาก PDF (ใช้ตรวจสอบเมื่อค่าไม่ถูกต้อง)"):
        addendum_match = re.search(
            r'(?:ADDENDUM OF THE CONTRACT 1|บันทึกแนบท้ายสัญญา\s*1).*',
            full_text, re.DOTALL | re.IGNORECASE)
        snippet = addendum_match.group(0)[:3000] if addendum_match else full_text[:3000]
        st.text(snippet)

    st.info("⚠️ กรุณาตรวจสอบค่าด้านบนก่อนกด Generate — ถ้าไม่ถูกต้องสามารถแก้ไขได้ด้านล่าง")

    with st.expander("✏️ แก้ไข fee rates (ถ้าจำเป็น)"):
        st.caption("ใส่เป็นตัวเลข % เช่น 20 หมายถึง 20%")
        for key, label in [('fixed','Fixed income'), ('variable','Variable income'),
                            ('expense','Expense/Reimbursement'), ('pvd','Provident Fund'),
                            ('health_ins','Health Insurance'), ('compensation','Severance/Compensation')]:
            if key in prtr_cats:
                st.write(f"**{label}**: PRTR (ตามสัญญา)")
                continue
            current = pad4(rates.get(key, [0.0]))
            edit_cols = st.columns(n_cols)
            new_vals = []
            for i, (c, lbl) in enumerate(zip(edit_cols, col_labels_active)):
                v = current[i] if i < len(current) else current[-1]
                nv = c.number_input(f"{label} - {lbl}", value=round(v*100, 2),
                                    min_value=0.0, max_value=100.0, step=0.5,
                                    key=f"{key}_{i}") / 100
                new_vals.append(nv)
            rates[key] = pad4(new_vals)

    st.divider()
    if st.button("🚀 Generate Fee Table", type="primary", use_container_width=True):
        with st.spinner("กำลังสร้างไฟล์..."):
            xlsx_file.seek(0)
            out_bytes = fill_excel(xlsx_file.read(), rates, n_cols, prtr_cats)
            output_filename = f"Master - Fee tebla_{company_name}.xlsx"
        st.success("✅ สร้างไฟล์เสร็จแล้ว!")
        st.download_button(
            label=f"⬇️ ดาวน์โหลด {output_filename}",
            data=out_bytes,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

else:
    st.info("👆 กรุณาอัปโหลดทั้ง 2 ไฟล์ด้านบนเพื่อเริ่มต้น")

st.divider()
st.caption("PRTR Group Public Company Limited · Fee Table Automation Tool")
